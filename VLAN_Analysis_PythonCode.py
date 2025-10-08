#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, re, json, argparse
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def read_text(p):
    with open(p, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def devname(text, fallback):
    m = re.search(r"<([A-Za-z0-9\-\._]+)>", text)
    if m:
        return m.group(1).split(".")[0]
    m = re.search(r"(?mi)^\s*sysname\s+([A-Za-z0-9\-_\.]+)", text)
    if m:
        return m.group(1).split(".")[0]
    return os.path.splitext(os.path.basename(fallback))[0]

def expand_vlan_tokens(expr):
    vids = set()
    expr = re.sub(r"#.*", "", expr)
    toks = re.split(r"[,\s]+", expr.strip())
    i = 0
    while i < len(toks):
        t = toks[i]
        if re.fullmatch(r"\d+", t):
            vids.add(int(t)); i += 1
        elif re.fullmatch(r"\d+\-\d+", t):
            a,b = map(int, t.split("-"))
            if a <= b: vids.update(range(a, b+1))
            i += 1
        elif t.lower() == "to" and i-1 >= 0 and i+1 < len(toks) and toks[i-1].isdigit() and toks[i+1].isdigit():
            a,b = int(toks[i-1]), int(toks[i+1])
            if a <= b: vids.update(range(a, b+1))
            i += 2
        else:
            i += 1
    return vids

def parse_declared(text):
    singles = set(); ranges = []
    for m in re.finditer(r"(?:The VLANs include|The following VLANs exist):\s*(.+)", text):
        chunk = m.group(1)
        tail = text[m.end():].splitlines()
        for ln in tail[:40]:
            if not ln.strip(): break
            if ln.strip().startswith("<"): break
            chunk += " " + ln.strip()
        chunk = chunk.replace("(default)", " ")
        for tok in re.split(r"[,\s]+", chunk.strip()):
            if re.fullmatch(r"\d+", tok):
                singles.add(int(tok))
            elif re.fullmatch(r"\d+\-\d+", tok):
                a,b = map(int, tok.split("-"))
                if a <= b: ranges.append((a,b))
    for m in re.finditer(r"(?mi)^vlan\s+(\d+)\s+to\s+(\d+)", text):
        a,b = int(m.group(1)), int(m.group(2))
        if a <= b: ranges.append((a,b))
    for m in re.finditer(r"(?mi)^vlan\s+(\d+)\-(\d+)", text):
        a,b = int(m.group(1)), int(m.group(2))
        if a <= b: ranges.append((a,b))
    for m in re.finditer(r"(?mi)^vlan\s+(\d+)\s*(?:\n|$)", text):
        singles.add(int(m.group(1)))
    return singles, ranges

def parse_vlan_names(text):
    names = {}
    for m in re.finditer(r"(?ms)^vlan\s+(\d+)(?:\s+to\s+\d+)?\s*\n(.*?)(?=^\S|\Z)", text):
        vid = int(m.group(1)); block = m.group(2)
        nm = re.search(r"(?mi)^\s*name\s+(.+)$", block)
        if nm:
            names[vid] = nm.group(1).strip()
        else:
            dm = re.search(r"(?mi)^\s*description\s+(.+)$", block)
            if dm: names[vid] = dm.group(1).strip()
    return names

def parse_svi(text):
    svi = defaultdict(set)
    for m in re.finditer(r"(?ms)^interface\s+Vlan(?:-interface)?(\d+)\s*\n(.*?)(?=^interface|\Z)", text):
        vid = int(m.group(1)); block = m.group(2)
        for ipm in re.finditer(r"(?mi)ip address\s+([0-9\.]+)\s+[0-9\.]+", block):
            svi[vid].add(ipm.group(1))
    for m in re.finditer(r"(?mi)^Vlan(?:-interface)?\s*(\d+)\s+\S+\s+\S+\s+([0-9\.]+)", text):
        svi[int(m.group(1))].add(m.group(2))
    return {k: sorted(v) for k,v in svi.items()}

def parse_access(text):
    vids = set()
    for m in re.finditer(r"(?mi)port\s+(?:access|default)\s+vlan\s+(\d+)", text):
        vids.add(int(m.group(1)))
    for m in re.finditer(r"(?mi)port\s+hybrid\s+pvid\s+vlan\s+(\d+)", text):
        vids.add(int(m.group(1)))
    for m in re.finditer(r"(?mi)port\s+hybrid\s+untagged\s+vlan\s+([0-9\-\s,]+)", text):
        vids |= expand_vlan_tokens(m.group(1))
    bt = re.search(r"Interface\s+Link\s+Speed.*?Type\s+PVID.*?\n(.+?)(?=\n\s*\n|<|\Z)", text, re.S)
    if bt:
        for ln in bt.group(1).splitlines():
            cols = re.split(r"\\s{2,}", ln.strip())
            if len(cols) < 6: 
                continue
            typev = None
            for c in cols:
                if c in ("A","H","T"):
                    typev = c; break
            if typev == "T":
                continue
            pvid = None
            for c in reversed(cols):
                if re.fullmatch(r"\\d+", c):
                    pvid = int(c); break
            if pvid is not None:
                vids.add(pvid)
    return vids

def parse_trunk_rules(text):
    results = {}
    for m in re.finditer(r"(?ms)^interface\\s+([A-Za-z0-9/\\-]+)\\s*\\n(.*?)(?=^interface|\\Z)", text):
        iface = m.group(1); body = m.group(2)
        all_flag = False; exceptions = set(); allow_list = set(); hybrid_tagged = set()
        is_trunk = bool(re.search(r"(?mi)port\\s+link-type\\s+trunk", body))
        if is_trunk:
            for mm in re.finditer(r"(?mi)port\\s+trunk\\s+permit\\s+vlan\\s+([^\\n#]+)", body):
                expr = mm.group(1).strip()
                if expr.lower().startswith("all") or re.search(r"\\b2\\s+to\\s+4094\\b", expr):
                    all_flag = True
                else:
                    allow_list |= expand_vlan_tokens(expr)
            for um in re.finditer(r"(?mi)undo\\s+port\\s+trunk\\s+permit\\s+vlan\\s+([^\\n#]+)", body):
                exceptions |= expand_vlan_tokens(um.group(1).strip())
        for mm in re.finditer(r"(?mi)port\\s+hybrid\\s+tagged\\s+vlan\\s+([^\\n#]+)", body):
            hybrid_tagged |= expand_vlan_tokens(mm.group(1).strip())
        if is_trunk or hybrid_tagged:
            results[iface] = {"all": all_flag, "except": exceptions, "list": allow_list, "hybrid": hybrid_tagged}
    return results

def role_for(dev, core_names=None, agg_prefix=None, role_map=None):
    dn = dev.lower()
    if role_map and dev in role_map:
        return role_map[dev]
    if core_names and dn in core_names:
        return "Core"
    if agg_prefix and agg_prefix in dn:
        return "Aggregation"
    return "Access"

def merge_intervals(intervals):
    if not intervals:
        return []
    intervals = sorted(intervals)
    merged = [intervals[0]]
    for s,e in intervals[1:]:
        ls, le = merged[-1]
        if s <= le + 1:
            merged[-1] = (ls, max(le, e))
        else:
            merged.append((s, e))
    return merged

def points_to_segments(points):
    if not points:
        return []
    pts = sorted(set(points))
    segs = []
    start = prev = pts[0]
    for p in pts[1:]:
        if p == prev + 1:
            prev = p
        else:
            segs.append((start, prev))
            start = prev = p
    segs.append((start, prev))
    return segs

def subtract_points(intervals, points):
    pts = sorted(set(points))
    out = []
    for s,e in intervals:
        cur = s
        for p in pts:
            if p < s or p > e:
                continue
            if p == cur:
                cur = p + 1
            else:
                out.append((cur, p-1))
                cur = p + 1
        if cur <= e:
            out.append((cur, e))
    return out

def audit_folder(input_dir, output_path, core_names=None, agg_prefix="espac", role_map=None):
    devices = {}
    for name in sorted(os.listdir(input_dir)):
        if os.path.splitext(name)[1].lower() not in (".cfg", ".log", ".txt"):
            continue
        p = os.path.join(input_dir, name)
        txt = read_text(p)
        if not txt.strip():
            continue
        dev = devname(txt, p)
        info = devices.setdefault(dev, {
            "declared_singles": set(),
            "declared_ranges": [],
            "vlan_names": {},
            "svi": defaultdict(list),
            "access": set(),
            "trunk_ifaces": {},
        })
        singles, ranges = parse_declared(txt)
        info["declared_singles"] |= singles
        info["declared_ranges"].extend(ranges)
        info["vlan_names"].update(parse_vlan_names(txt))
        svi = parse_svi(txt)
        for vid, ips in svi.items():
            cur = set(info["svi"].get(vid, [])); cur.update(ips)
            info["svi"][vid] = sorted(cur)
        info["access"] |= parse_access(txt)
        info["trunk_ifaces"].update(parse_trunk_rules(txt))

    all_devices = sorted(devices.keys())

    dev_all_trunk = {dev: any(r["all"] for r in d["trunk_ifaces"].values()) for dev, d in devices.items()}
    dev_explicit_trunk = {
        dev: set().union(*[(r["list"] | r["hybrid"]) for r in d["trunk_ifaces"].values()]) if d["trunk_ifaces"] else set()
        for dev, d in devices.items()
    }

    used_by_hosts = sorted(set().union(*[ set(d["svi"].keys()) | set(d["access"]) for d in devices.values() ]))
    exp_trunk_only_points = sorted(set().union(*[dev_explicit_trunk[dev] for dev in all_devices]) - set(used_by_hosts))
    merged_declared_ranges = merge_intervals([rng for d in devices.values() for rng in d["declared_ranges"]])

    unused_segments = subtract_points(merged_declared_ranges, used_by_hosts + exp_trunk_only_points)
    exp_trunk_segments = points_to_segments(exp_trunk_only_points)

    rows = []
    any_trunk_all_devices = [dev for dev in all_devices if dev_all_trunk.get(dev, False)]

    for (s,e) in unused_segments:
        decl_devs = []
        for dev, d in devices.items():
            for rs,re in merge_intervals(d["declared_ranges"]):
                if not (e < rs or s > re):
                    decl_devs.append(dev); break
        color = "Yellow" if len(any_trunk_all_devices) > 0 else "Red"
        usage = "Verify" if color == "Yellow" else "Delete_Candidate"
        rows.append({
            "VLAN_ID": f"{s}-{e}" if s != e else f"{s}",
            "Color": color,
            "Usage_Category": usage,
            "Description": f"Declared by range {s}-{e}; unused; {'transit via trunk ALL possible' if color=='Yellow' else 'no usage observed'}",
            "Devices_Total": len(set(decl_devs + any_trunk_all_devices)),
            "Devices_Core": sum(1 for dev in set(decl_devs + any_trunk_all_devices) if role_for(dev, core_names, agg_prefix, role_map)=="Core"),
            "Devices_Aggregation": sum(1 for dev in set(decl_devs + any_trunk_all_devices) if role_for(dev, core_names, agg_prefix, role_map)=="Aggregation"),
            "Devices_Access": sum(1 for dev in set(decl_devs + any_trunk_all_devices) if role_for(dev, core_names, agg_prefix, role_map)=="Access"),
            "Devices_with_SVI": "",
            "Devices_with_Access(PVID/Untagged)": "",
            "Devices_with_Tagged_Trunk": ", ".join(sorted(any_trunk_all_devices)),
            "SVI_IPs": "",
            "Devices_with_VLAN_Config(Seen)": ", ".join(sorted(set(decl_devs))),
        })

    for (s,e) in exp_trunk_segments:
        tag_explicit_devs = []
        for dev in all_devices:
            if any(v in dev_explicit_trunk[dev] for v in range(s, e+1)):
                tag_explicit_devs.append(dev)
        rows.append({
            "VLAN_ID": f"{s}-{e}" if s != e else f"{s}",
            "Color": "Yellow",
            "Usage_Category": "Verify",
            "Description": f"Transit only (explicit trunk), {s}-{e}",
            "Devices_Total": len(set(tag_explicit_devs)),
            "Devices_Core": sum(1 for dev in set(tag_explicit_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Core"),
            "Devices_Aggregation": sum(1 for dev in set(tag_explicit_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Aggregation"),
            "Devices_Access": sum(1 for dev in set(tag_explicit_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Access"),
            "Devices_with_SVI": "",
            "Devices_with_Access(PVID/Untagged)": "",
            "Devices_with_Tagged_Trunk": ", ".join(sorted(tag_explicit_devs)),
            "SVI_IPs": "",
            "Devices_with_VLAN_Config(Seen)": "",
        })

    for vid in used_by_hosts:
        svi_devs = sorted([dev for dev in all_devices if vid in devices[dev]["svi"]])
        acc_devs = sorted([dev for dev in all_devices if vid in devices[dev]["access"]])
        tag_devs = sorted([dev for dev in all_devices if (vid in dev_explicit_trunk[dev]) or dev_all_trunk[dev]])
        name_candidates = [devices[dev]["vlan_names"].get(vid, "") for dev in all_devices if vid in devices[dev]["vlan_names"]]
        name_candidates = [n for n in name_candidates if n]
        desc = name_candidates[0] if name_candidates else "Unknown"
        color = "Green" if desc != "Unknown" else "White"
        rows.append({
            "VLAN_ID": vid,
            "Color": color,
            "Usage_Category": "Described" if color=="Green" else "In_Use",
            "Description": desc,
            "Devices_Total": len(set(svi_devs+acc_devs+tag_devs)),
            "Devices_Core": sum(1 for dev in set(svi_devs+acc_devs+tag_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Core"),
            "Devices_Aggregation": sum(1 for dev in set(svi_devs+acc_devs+tag_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Aggregation"),
            "Devices_Access": sum(1 for dev in set(svi_devs+acc_devs+tag_devs) if role_for(dev, core_names, agg_prefix, role_map)=="Access"),
            "Devices_with_SVI": ", ".join(svi_devs),
            "Devices_with_Access(PVID/Untagged)": ", ".join(acc_devs),
            "Devices_with_Tagged_Trunk": ", ".join(tag_devs),
            "SVI_IPs": ", ".join(sorted({ip for dev in svi_devs for ip in devices[dev]["svi"][vid]})),
            "Devices_with_VLAN_Config(Seen)": ", ".join(sorted([dev for dev in all_devices if any(s<=vid<=e for s,e in devices[dev]["declared_ranges"])])),
        })

    pvg_df = pd.DataFrame(rows).sort_values(["Color","VLAN_ID"], key=lambda s: s.astype(str)).reset_index(drop=True)

    pds_rows = []
    for dev in all_devices:
        d = devices[dev]
        seen = set(d["svi"].keys()) | set(d["access"]) | set(dev_explicit_trunk[dev])
        is_all = dev_all_trunk[dev]
        count = "ALL" if is_all else (len(dev_explicit_trunk[dev]) if len(dev_explicit_trunk[dev])>0 else 0)
        if isinstance(count, int) and count > 2000:
            count = "ALL"
        pds_rows.append({
            "Device": dev,
            "Role": role_for(dev, core_names, agg_prefix, role_map),
            "VLANs_Defined": len(seen),
            "SVI_Count": len(d["svi"]),
            "Tagged_Trunk_VLANs(Count)": count,
        })
    pds_df = pd.DataFrame(pds_rows).sort_values(["Role","Device"])

    pdvd_rows = []
    for dev in all_devices:
        d = devices[dev]
        seen = sorted(set(d["svi"].keys()) | set(d["access"]) | set(dev_explicit_trunk[dev]))
        for vid in seen:
            pdvd_rows.append({
                "Device": dev,
                "Role": role_for(dev, core_names, agg_prefix, role_map),
                "VLAN_ID": vid,
                "SVI?": "Yes" if vid in d["svi"] else "No",
                "SVI_IPs": ", ".join(d["svi"].get(vid, [])),
                "Tagged on Trunk/Hybrid?": "Yes" if (vid in dev_explicit_trunk[dev]) or dev_all_trunk[dev] else "No",
                "VLAN_Name": devices[dev]["vlan_names"].get(vid, ""),
            })
    pdvd_df = pd.DataFrame(pdvd_rows).sort_values(["Device","VLAN_ID"])

    rows_matrix = []
    for vid in used_by_hosts:
        row = {"VLAN_ID": vid}
        for dev in all_devices:
            d = devices[dev]
            flags = []
            if vid in d["svi"]:
                flags.append("S")
            if vid in d["access"]:
                flags.append("A")
            if (vid in dev_explicit_trunk[dev]) or dev_all_trunk[dev]:
                flags.append("T")
            row[f"{dev} [{role_for(dev, core_names, agg_prefix, role_map)}]"] = "+".join(flags) if flags else None
        rows_matrix.append(row)
    matrix_df = pd.DataFrame(rows_matrix)

    roles_df = pd.DataFrame([{"Device": dev, "Role": role_for(dev, core_names, agg_prefix, role_map)} for dev in all_devices])
    summary_df = pd.DataFrame([{
        "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Devices_Parsed": len(all_devices),
        "Per_VLAN_Global_Rows": len(pvg_df),
        "Green(Described)": int((pvg_df["Color"]=="Green").sum()),
        "White(In_Use)": int((pvg_df["Color"]=="White").sum()),
        "Yellow(Verify)": int((pvg_df["Color"]=="Yellow").sum()),
        "Red(Delete_Candidate)": int((pvg_df["Color"]=="Red").sum()),
    }])

    explain_df = pd.DataFrame({
        "Summary": [
            "VLAN audit per your CSW template.",
            "In-use VLANs (SVI/Access) listed individually; transit-only & unused-declared are grouped as contiguous ranges.",
            "Color rules: Green=has name/description; White=in-use w/o name; Yellow=transit-only or verify (e.g., trunk ALL present); Red=delete candidate (no SVI/Access/explicit trunk and no trunk ALL).",
            "Tagged_Trunk_VLANs(Count)=ALL if any port has 'port trunk permit vlan all' or 'permit vlan 2 to 4094'. 'undo port trunk permit vlan 1' excludes VLAN 1 from that port."
        ],
        "Notes": [f"Input folder: {input_dir}", "", "", ""],
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        explain_df.to_excel(writer, index=False, sheet_name="Explain")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        roles_df.to_excel(writer, index=False, sheet_name="Roles")
        pvg_df.to_excel(writer, index=False, sheet_name="Per_VLAN_Global(Clarity)")
        pds_df.to_excel(writer, index=False, sheet_name="Per_Device_Summary")
        pdvd_df.to_excel(writer, index=False, sheet_name="Per_Device_VLAN_Detail")
        matrix_df.to_excel(writer, index=False, sheet_name="VLAN x Device Matrix")

    wb = load_workbook(output_path)
    ws = wb["Per_VLAN_Global(Clarity)"]
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    color_col = header.get("Color", 2)
    fills = {
        "Green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "White": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "Yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    for r in range(2, ws.max_row+1):
        val = ws.cell(row=r, column=color_col).value
        if val in fills:
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = fills[val]
    wb.save(output_path)

def main():
    parser = argparse.ArgumentParser(description="H3C VLAN Audit -> Excel (grouped + colored).")
    parser.add_argument("--input", required=True, help="Folder containing H3C logs/configs (.cfg/.log/.txt)")
    parser.add_argument("--output", required=True, help="Output Excel path")
    parser.add_argument("--core-names", nargs="*", default=["espcsw03"], help="Device names treated as Core (default: espcsw03)")
    parser.add_argument("--agg-prefix", default="espac", help="Substring to mark device as Aggregation (default: 'espac')")
    parser.add_argument("--role-map", default=None, help="JSON file mapping device->role (Core/Aggregation/Access)")
    args = parser.parse_args()

    role_map = None
    if args.role_map and os.path.exists(args.role_map):
        with open(args.role_map, "r", encoding="utf-8") as f:
            role_map = json.load(f)

    res = audit_folder(args.input, args.output, core_names=set([n.lower() for n in args.core_names]), agg_prefix=args.agg_prefix.lower(), role_map=role_map)

if __name__ == "__main__":
    main()
